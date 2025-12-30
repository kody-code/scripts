from openpyxl import load_workbook

def read_excel(file_path):
    wb = load_workbook(file_path, read_only=True)
    sheet_name = "Sheet1"
    if sheet_name not in wb.sheetnames:
        print("Sheet not found")
        exit(1)
    sheet = wb[sheet_name]

    # 读取特定单元格的数据
    cell_value = sheet['A2'].value

    all_data = []
    for row_num in range(2, sheet.max_row + 1):
        row_data = []
        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            row_data.append(cell_value)
        all_data.append(row_data)

    wb.close()

    return all_data

# if __name__ == '__main__':
#     data = read_excel("../doc/申请发票.xlsx")
#     for row in data:
#         print(f"合同号：{row[0]}, 金额： {row[1]}, 发票内容： {row[2]}")